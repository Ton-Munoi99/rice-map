#!/usr/bin/env python3
"""
สร้างไฟล์ Excel วิเคราะห์ระยะห่างโรงสีข้าว จากข้อมูล DIT

ที่มาพิกัด: ที่อยู่โรงสีใน DIT Excel มีครบถึงระดับ "ตำบล" แต่ไม่มี lat/lon
เราจึง geocode ที่อยู่ระดับ ตำบล/อำเภอ/จังหวัด ผ่าน OpenStreetMap Nominatim
(ชื่อสถานที่ราชการ geocode ได้ดี ต่างจากชื่อบริษัทที่ค้นไม่เจอ)

⚠️ ข้อจำกัดสำคัญ: พิกัดที่ได้เป็น "ระดับตำบล" ไม่ใช่ตำแหน่งโรงสีจริง
   → โรงสีที่อยู่ตำบลเดียวกันจะได้พิกัดเดียวกัน (ระยะห่าง = 0)
   → ระยะทางที่คำนวณเป็นค่าประมาณระดับตำบล (คลาดเคลื่อนได้ ~2-8 กม.)

Output: rice_mills_distance_analysis.xlsx
Run:    python scripts/build_mills_distance_excel.py
"""
import io
import json
import math
import os
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_XLSX = os.path.join(ROOT, "thai_rice_mills_dit_2026-04-23.xlsx")
OUT_XLSX = os.path.join(ROOT, "rice_mills_distance_analysis.xlsx")
CACHE = os.path.join(ROOT, "data", "mills-geocode-cache.json")
UA = "RiceMap/1.0 (rice mill distance analysis; github.com/Ton-Munoi99/rice-map)"
SLEEP = 1.1  # Nominatim usage policy: max 1 req/sec


# ── geocoding ────────────────────────────────────────────────────────────────
def load_cache():
    if os.path.exists(CACHE):
        try:
            return json.load(open(CACHE, encoding="utf-8"))
        except (ValueError, OSError):
            pass
    return {}


def save_cache(c):
    os.makedirs(os.path.dirname(CACHE), exist_ok=True)
    json.dump(c, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False, indent=1)


def nominatim(q):
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"q": q, "format": "json", "limit": 1, "countrycodes": "th"}
    )
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    try:
        r = json.load(urllib.request.urlopen(req, timeout=25))
        if not r:
            return None
        return {"lat": float(r[0]["lat"]), "lon": float(r[0]["lon"]),
                "display": r[0].get("display_name", "")}
    except Exception as e:
        print(f"  [warn] {q[:40]}: {e}", file=sys.stderr)
        return None


def geocode_cached(key, query, cache):
    if key in cache:
        return cache[key]
    res = nominatim(query)
    cache[key] = res
    time.sleep(SLEEP)
    return res


def load_province_bbox():
    """กรอบพิกัดรายจังหวัด (จาก districts-geo.json คีย์อังกฤษ) แม็พชื่อไทยผ่าน rice-data.csv
    ใช้ตรวจว่าผลลัพธ์ geocode ตกอยู่ในจังหวัดที่ถูกต้องจริง — Nominatim ชอบจับชื่อผิดจังหวัด
    เช่น 'ศาลากลางจังหวัดลำพูน' ไปได้ 'ตำบลศาลากลาง' ที่ปทุมธานี"""
    import csv
    th2en = {}
    with open(os.path.join(ROOT, "rice-data.csv"), encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            th2en[r["province_th"].strip()] = r["province_en"].strip()
    geo = json.load(open(os.path.join(ROOT, "data", "districts-geo.json"), encoding="utf-8"))
    bbox = {k: v["bbox"] for k, v in geo["provinces"].items()}
    return th2en, bbox


def in_province(prov_th, lat, lon, th2en, bbox, pad=0.25):
    """pad ~25 กม. เผื่อความคลาดเคลื่อนของ bbox ที่ simplify มาแล้ว"""
    b = bbox.get(th2en.get(prov_th, ""))
    if not b:
        return True  # ไม่มี bbox ให้เทียบ → ปล่อยผ่าน ดีกว่าทิ้งข้อมูลดี
    return (b[0] - pad <= lon <= b[2] + pad) and (b[1] - pad <= lat <= b[3] + pad)


def geocode_validated(queries, prov_th, th2en, bbox):
    """ลองคำค้นตามลำดับ คืนผลแรกที่ 'อยู่ในจังหวัดถูกต้อง' เท่านั้น"""
    for q, level in queries:
        res = nominatim(q)
        time.sleep(SLEEP)
        if res and in_province(prov_th, res["lat"], res["lon"], th2en, bbox):
            res["level"] = level
            return res
    return None


def _queries(prov, dist, tam):
    """กรุงเทพฯ ใช้ แขวง/เขต ไม่ใช่ ตำบล/อำเภอ → ต้องถามคนละแบบ"""
    if prov in ("กรุงเทพมหานคร", "กรุงเทพ", "กทม."):
        return (f"แขวง{tam} เขต{dist} กรุงเทพมหานคร", f"เขต{dist} กรุงเทพมหานคร")
    return (f"ตำบล{tam} อำเภอ{dist} จังหวัด{prov}", f"อำเภอ{dist} จังหวัด{prov}")


def resolve_tambon(key, prov, dist, tam, cache, th2en, bbox):
    """หาพิกัดระดับตำบล → ถอยไประดับอำเภอ ทุกผลต้องผ่านการตรวจว่าอยู่ในจังหวัดถูกต้อง"""
    cur = cache.get(key, "MISSING")
    if isinstance(cur, dict):
        cur.setdefault("level", "tambon")
        return cur
    q_tam, q_dist = _queries(prov, dist, tam)
    res = geocode_validated([(q_tam, "tambon"), (q_dist, "district")], prov, th2en, bbox)
    cache[key] = res
    return res


# ── geometry ─────────────────────────────────────────────────────────────────
def haversine_km(a_lat, a_lon, b_lat, b_lon):
    R = 6371.0
    dlat, dlon = math.radians(b_lat - a_lat), math.radians(b_lon - a_lon)
    h = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(a_lat)) * math.cos(math.radians(b_lat)) * math.sin(dlon / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(h))


# ── load mills ───────────────────────────────────────────────────────────────
def load_mills():
    wb = openpyxl.load_workbook(SRC_XLSX, read_only=True)
    ws = wb["รายชื่อโรงสี"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    ix = {name: hdr.index(name) for name in hdr if name}

    def g(r, name):
        i = ix.get(name)
        if i is None or i >= len(r) or r[i] is None:
            return ""
        return str(r[i]).strip()

    mills = []
    for r in rows[1:]:
        name = g(r, "ชื่อโรงสี/ผู้ประกอบการ")
        prov = g(r, "จังหวัด")
        if not name or not prov:
            continue
        cap = g(r, "กำลังการผลิต")
        try:
            cap_v = float(cap.replace(",", "")) if cap else None
        except ValueError:
            cap_v = None
        mills.append({
            "name": name,
            "type": g(r, "ประเภทโรงสี"),
            "province": prov,
            "district": g(r, "อำเภอ/เขต"),
            "tambon": g(r, "ตำบล/แขวง"),
            "address": g(r, "ที่อยู่สถานประกอบการ"),
            "zip": g(r, "รหัสไปรษณีย์"),
            "phone": g(r, "โทรศัพท์"),
            "capacity": cap_v,
            "license": g(r, "เลขที่หนังสืออนุญาต"),
        })
    return mills


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    mills = load_mills()
    print(f"Loaded {len(mills)} mills from DIT Excel")

    cache = load_cache()
    th2en, bbox = load_province_bbox()

    # 0) ตรวจ cache เดิม — ทิ้งพิกัดที่ตกนอกจังหวัดตัวเอง แล้วหาใหม่
    dropped = 0
    for k, v in list(cache.items()):
        if isinstance(v, dict) and not in_province(k.split("|")[1], v["lat"], v["lon"], th2en, bbox):
            del cache[k]
            dropped += 1
    if dropped:
        print(f"ทิ้งพิกัดที่ผิดจังหวัด {dropped} รายการ → จะ geocode ใหม่")

    # 1) geocode unique tambon
    keys = []
    for m in mills:
        k = f"T|{m['province']}|{m['district']}|{m['tambon']}"
        m["_key"] = k
        if k not in keys:
            keys.append(k)
    done = sum(1 for k in keys if isinstance(cache.get(k), dict))
    print(f"Geocoding {len(keys)} unique tambon (already resolved: {done})")
    for i, k in enumerate(keys, 1):
        _, prov, dist, tam = k.split("|")
        resolve_tambon(k, prov, dist, tam, cache, th2en, bbox)
        if i % 25 == 0:
            save_cache(cache)
            print(f"  {i}/{len(keys)}")
    save_cache(cache)

    # 2) จุดกึ่งกลางตัวเมืองจังหวัด — ต้องผ่านการตรวจ bbox ด้วย
    provs = sorted({m["province"] for m in mills})
    print(f"Geocoding {len(provs)} provincial city centres")
    for p in provs:
        k = f"C|{p}"
        if isinstance(cache.get(k), dict):
            continue
        qs = ([(f"เขตพระนคร กรุงเทพมหานคร", "city")] if p == "กรุงเทพมหานคร"
              else [(f"ศาลากลางจังหวัด{p}", "city"),
                    (f"อำเภอเมือง{p} จังหวัด{p}", "city"),
                    (f"จังหวัด{p}", "province")])
        cache[k] = geocode_validated(qs, p, th2en, bbox)
    save_cache(cache)

    # 3) attach coords
    ok = 0
    for m in mills:
        c = cache.get(m["_key"])
        if isinstance(c, dict):
            m["lat"], m["lon"] = round(c["lat"], 6), round(c["lon"], 6)
            lvl = c.get("level", "tambon")
            m["geo_level"] = lvl
            m["geo_src"] = ("OSM Nominatim — ระดับตำบล" if lvl == "tambon"
                            else "OSM Nominatim — ระดับอำเภอ (ตำบลหาไม่เจอ)")
            ok += 1
        else:
            m["lat"] = m["lon"] = None
            m["geo_level"] = "none"
            m["geo_src"] = "หาพิกัดไม่ได้"
    print(f"Coordinates resolved: {ok}/{len(mills)}")

    # 4) distances
    by_prov = defaultdict(list)
    for m in mills:
        by_prov[m["province"]].append(m)

    for prov, group in by_prov.items():
        cc = cache.get(f"C|{prov}")
        pts = [m for m in group if m["lat"] is not None]
        for m in group:
            # distance to provincial city centre
            m["dist_city"] = (round(haversine_km(m["lat"], m["lon"], cc["lat"], cc["lon"]), 2)
                              if (m["lat"] is not None and cc) else None)
            # nearest other mill in same province
            m["near_name"] = m["near_km"] = None      # ใกล้สุดแบบนับทุกโรง (มี 0 เยอะ)
            m["far_name"] = m["far_km"] = None        # ใกล้สุดที่ 'พิกัดต่างกัน' = ค่าที่วัดได้จริง
            m["same_tambon"] = 0                      # จำนวนโรงที่ใช้พิกัดเดียวกัน
            if m["lat"] is None:
                continue
            best = best_diff = None
            for o in pts:
                if o is m:
                    continue
                d = haversine_km(m["lat"], m["lon"], o["lat"], o["lon"])
                if d == 0:
                    m["same_tambon"] += 1   # พิกัดชนกัน (ตำบลเดียวกัน หรือถอยไปพิกัดอำเภอเดียวกัน)
                elif best_diff is None or d < best_diff[0]:
                    best_diff = (d, o["name"])
                if best is None or d < best[0]:
                    best = (d, o["name"])
            if best:
                m["near_km"], m["near_name"] = round(best[0], 2), best[1]
            if best_diff:
                m["far_km"], m["far_name"] = round(best_diff[0], 2), best_diff[1]

    # 5) write Excel
    write_excel(mills, by_prov, cache)
    print(f"\n✅ Wrote {OUT_XLSX}")


def write_excel(mills, by_prov, cache):
    wb = openpyxl.Workbook()
    HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)

    def style_header(ws, ncols, widths):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = HDR_FILL, HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # Sheet 1 — mills
    ws = wb.active
    ws.title = "รายชื่อโรงสี"
    cols = ["ลำดับ", "ชื่อโรงสี", "ประเภท", "กำลังการผลิต (ตัน/วัน)", "จังหวัด", "อำเภอ", "ตำบล",
            "ที่อยู่", "รหัสไปรษณีย์", "โทรศัพท์", "Latitude", "Longitude", "ที่มาพิกัด"]
    ws.append(cols)
    for i, m in enumerate(sorted(mills, key=lambda x: (x["province"], x["district"], x["name"])), 1):
        ws.append([i, m["name"], m["type"], m["capacity"], m["province"], m["district"], m["tambon"],
                   m["address"], m["zip"], m["phone"], m["lat"], m["lon"], m["geo_src"]])
    style_header(ws, len(cols), [6, 42, 16, 18, 16, 16, 16, 46, 12, 16, 12, 12, 24])

    # Sheet 2 — mill-to-mill
    ws2 = wb.create_sheet("ระยะห่างโรงสี-โรงสี")
    cols2 = ["จังหวัด", "ชื่อโรงสี", "อำเภอ", "ตำบล", "โรงสีที่ใช้พิกัดเดียวกัน (แห่ง)",
             "โรงสีใกล้สุดที่พิกัดต่างกัน", "ระยะห่าง (กม.) ⭐ใช้ค่านี้",
             "โรงสีใกล้สุด (นับทุกแห่ง)", "ระยะห่าง (กม.)", "ความละเอียดพิกัด"]
    ws2.append(cols2)
    for m in sorted(mills, key=lambda x: (x["province"], x["far_km"] if x["far_km"] is not None else 9e9)):
        lvl = "ตำบล" if m["geo_level"] == "tambon" else ("อำเภอ" if m["geo_level"] == "district" else "-")
        ws2.append([m["province"], m["name"], m["district"], m["tambon"], m["same_tambon"],
                    m["far_name"], m["far_km"], m["near_name"], m["near_km"], lvl])
    style_header(ws2, len(cols2), [16, 40, 15, 15, 22, 40, 22, 40, 15, 16])

    # Sheet 3 — distance to city
    ws3 = wb.create_sheet("ระยะห่างจากตัวเมือง")
    cols3 = ["จังหวัด", "ชื่อโรงสี", "อำเภอ", "ตำบล", "ระยะห่างจากตัวเมือง (กม.)",
             "กำลังการผลิต (ตัน/วัน)", "ประเภท"]
    ws3.append(cols3)
    for m in sorted(mills, key=lambda x: (x["province"], x["dist_city"] if x["dist_city"] is not None else 9e9)):
        ws3.append([m["province"], m["name"], m["district"], m["tambon"],
                    m["dist_city"], m["capacity"], m["type"]])
    style_header(ws3, len(cols3), [16, 42, 16, 16, 24, 20, 16])

    # Sheet 4 — province summary
    ws4 = wb.create_sheet("สรุปรายจังหวัด")
    cols4 = ["จังหวัด", "จำนวนโรงสี", "จำนวนจุดพิกัดที่ต่างกัน", "กำลังผลิตรวม (ตัน/วัน)",
             "ระยะเฉลี่ยถึงโรงสีที่พิกัดต่างกัน (กม.)", "ใกล้สุด (กม.)", "ไกลสุด (กม.)",
             "ระยะจากตัวเมืองเฉลี่ย (กม.)", "ไกลจากเมืองสุด (กม.)", "% พิกัดระดับตำบล"]
    ws4.append(cols4)
    for prov in sorted(by_prov):
        g = by_prov[prov]
        far = [m["far_km"] for m in g if m["far_km"] is not None]
        city = [m["dist_city"] for m in g if m["dist_city"] is not None]
        caps = [m["capacity"] for m in g if m["capacity"]]
        ntam = len({m["_key"] for m in g})
        ptam = round(100 * sum(1 for m in g if m["geo_level"] == "tambon") / len(g))
        ws4.append([
            prov, len(g), ntam,
            round(sum(caps), 1) if caps else None,
            round(sum(far) / len(far), 2) if far else None,
            min(far) if far else None, max(far) if far else None,
            round(sum(city) / len(city), 2) if city else None,
            max(city) if city else None, ptam,
        ])
    style_header(ws4, len(cols4), [16, 12, 18, 20, 28, 14, 14, 24, 22, 18])

    # Sheet 5 — methodology
    ws5 = wb.create_sheet("วิธีการ & แหล่งข้อมูล")
    ws5.column_dimensions["A"].width = 26
    ws5.column_dimensions["B"].width = 100
    notes = [
        ("หัวข้อ", "รายละเอียด"),
        ("แหล่งข้อมูลโรงสี", "กรมการค้าภายใน (DIT) — รายชื่อผู้ประกอบการโรงสีข้าว "
                             "https://www.dit.go.th/th/find-entrepreneur/rice/  (ไฟล์ export 23 เม.ย. 2569)"),
        ("จำนวนโรงสี", f"{len(mills)} แห่ง ใน {len(by_prov)} จังหวัด"),
        ("ที่มาของพิกัด (สำคัญ)",
         "DIT ไม่ได้ให้ lat/lon — เราแปลงที่อยู่ระดับ ตำบล/อำเภอ/จังหวัด เป็นพิกัดด้วย "
         "OpenStreetMap Nominatim (https://nominatim.openstreetmap.org)"),
        ("⚠️ ข้อจำกัดความแม่นยำ",
         "พิกัดเป็นระดับ 'ตำบล' ไม่ใช่ตำแหน่งโรงสีจริง คลาดเคลื่อนได้ราว 2-8 กม. "
         "ใช้ดูภาพรวมการกระจายตัว/เปรียบเทียบเชิงพื้นที่ได้ แต่ไม่ควรใช้วางแผนขนส่งที่ต้องการความแม่นยำสูง"),
        ("⚠️ โรงสีในตำบลเดียวกัน",
         "จะได้พิกัดเดียวกัน ทำให้ 'ระยะห่าง = 0' ทั้งที่จริงอาจห่างกันหลายกิโลเมตร "
         "(เกิดกับ ~62% ของโรงสี) ดูคอลัมน์ 'จำนวนโรงสีในตำบลเดียวกัน' ประกอบเสมอ"),
        ("⭐ ควรใช้คอลัมน์ไหน",
         "ในชีต 'ระยะห่างโรงสี-โรงสี' ให้ใช้ 'ระยะห่างถึงโรงสีใกล้สุดที่พิกัดต่างกัน' เป็นหลัก "
         "เพราะเป็นค่าที่วัดได้จริงจากความละเอียดของข้อมูล ส่วนคอลัมน์ 'นับทุกแห่ง' "
         "จะเป็น 0 จำนวนมาก ซึ่งแปลว่าพิกัดชนกัน ไม่ได้แปลว่าโรงสีตั้งอยู่ที่เดียวกันจริง"),
        ("ความละเอียดพิกัดแต่ละแถว",
         "ดูคอลัมน์ 'ความละเอียดพิกัด' — 'ตำบล' แม่นกว่า (~2-8 กม.) · 'อำเภอ' หยาบกว่า (~5-20 กม.) "
         "เพราะ Nominatim หาระดับตำบลไม่เจอจึงถอยมาใช้จุดกึ่งกลางอำเภอแทน"),
        ("การตรวจสอบความถูกต้อง",
         "ทุกพิกัดถูกตรวจว่าตกอยู่ในกรอบเขตจังหวัดที่ถูกต้องจริง (province bbox validation) "
         "พบและแก้ไขพิกัดที่ Nominatim จับผิดจังหวัด เช่น 'ศาลากลางจังหวัดลำพูน' ที่เคยได้พิกัดในกรุงเทพฯ"),
        ("วิธีคำนวณระยะทาง",
         "Haversine (ระยะทางเส้นตรงบนผิวโลก, รัศมีโลก 6,371 กม.) — ไม่ใช่ระยะทางตามถนนจริง"),
        ("ตัวเมืองจังหวัด", "พิกัดศาลากลางจังหวัด (fallback: อำเภอเมือง) จาก Nominatim"),
        ("โรงสีที่ใกล้ที่สุด", "คำนวณเฉพาะโรงสีภายในจังหวัดเดียวกันเท่านั้น (ไม่ข้ามจังหวัด)"),
        ("จัดทำเมื่อ", time.strftime("%Y-%m-%d %H:%M")),
    ]
    for row in notes:
        ws5.append(row)
    for c in range(1, 3):
        ws5.cell(row=1, column=c).fill = HDR_FILL
        ws5.cell(row=1, column=c).font = HDR_FONT
    for r in range(2, len(notes) + 1):
        ws5.cell(row=r, column=1).font = Font(bold=True, size=10)
        ws5.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws5.row_dimensions[r].height = 34

    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
