#!/usr/bin/env python3
"""
Build data/flood-risk.json from DDPM (กรมป้องกันและบรรเทาสาธารณภัย) flood statistics.
Source: catalog.disaster.go.th/dataset/dpm-gd027
Period: 2562-2568 (7 years / 2019-2025)
Data: Official disaster declarations -- covers ALL flood types including flash floods
"""
import json, os, sys, io
from datetime import date

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, "..")
DDPM_DIR = os.path.join(ROOT, "data", "source", "ddpm")
OUTPUT = os.path.join(ROOT, "data", "flood-risk.json")

# ─────────────────────────────────────────────
# Province name mapping: Thai -> English
# (copied from fetch_miller_prices.py + additions)
# ─────────────────────────────────────────────
PROVINCE_MAP = {
    "กรุงเทพมหานคร": "Bangkok Metropolis",
    "กระบี่": "Krabi", "กาญจนบุรี": "Kanchanaburi",
    "กาฬสินธุ์": "Kalasin", "กำแพงเพชร": "Kamphaeng Phet",
    "ขอนแก่น": "Khon Kaen", "จันทบุรี": "Chanthaburi",
    "ฉะเชิงเทรา": "Chachoengsao", "ชลบุรี": "Chon Buri",
    "ชัยนาท": "Chai Nat", "ชัยภูมิ": "Chaiyaphum",
    "ชุมพร": "Chumphon", "เชียงราย": "Chiang Rai",
    "เชียงใหม่": "Chiang Mai", "ตรัง": "Trang",
    "ตราด": "Trat", "ตาก": "Tak",
    "นครนายก": "Nakhon Nayok", "นครปฐม": "Nakhon Pathom",
    "นครพนม": "Nakhon Phanom", "นครราชสีมา": "Nakhon Ratchasima",
    "นครศรีธรรมราช": "Nakhon Si Thammarat",
    "นครสวรรค์": "Nakhon Sawan", "นนทบุรี": "Nonthaburi",
    "นราธิวาส": "Narathiwat", "น่าน": "Nan",
    "บึงกาฬ": "Bueng Kan", "บุรีรัมย์": "Buri Ram",
    "ปทุมธานี": "Pathum Thani",
    "ประจวบคีรีขันธ์": "Prachuap Khiri Khan",
    "ปราจีนบุรี": "Prachin Buri", "ปัตตานี": "Pattani",
    "พระนครศรีอยุธยา": "Phra Nakhon Si Ayutthaya",
    "พะเยา": "Phayao", "พังงา": "Phangnga",
    "พัทลุง": "Phatthalung", "พิจิตร": "Phichit",
    "พิษณุโลก": "Phitsanulok", "เพชรบุรี": "Phetchaburi",
    "เพชรบูรณ์": "Phetchabun", "แพร่": "Phrae",
    "ภูเก็ต": "Phuket", "มหาสารคาม": "Maha Sarakham",
    "มุกดาหาร": "Mukdahan", "แม่ฮ่องสอน": "Mae Hong Son",
    "ยโสธร": "Yasothon", "ยะลา": "Yala",
    "ร้อยเอ็ด": "Roi Et", "ระนอง": "Ranong",
    "ระยอง": "Rayong", "ราชบุรี": "Ratchaburi",
    "ลพบุรี": "Lop Buri", "ลำปาง": "Lampang",
    "ลำพูน": "Lamphun", "เลย": "Loei",
    "ศรีสะเกษ": "Si Sa Ket", "สกลนคร": "Sakon Nakhon",
    "สงขลา": "Songkhla", "สตูล": "Satun",
    "สมุทรปราการ": "Samut Prakan",
    "สมุทรสงคราม": "Samut Songkhram",
    "สมุทรสาคร": "Samut Sakhon",
    "สระแก้ว": "Sa Kaeo", "สระบุรี": "Saraburi",
    "สิงห์บุรี": "Sing Buri", "สุโขทัย": "Sukhothai",
    "สุพรรณบุรี": "Suphan Buri", "สุราษฎร์ธานี": "Surat Thani",
    "สุรินทร์": "Surin", "หนองคาย": "Nong Khai",
    "หนองบัวลำภู": "Nong Bua Lam Phu",
    "อ่างทอง": "Ang Thong", "อำนาจเจริญ": "Amnat Charoen",
    "อุดรธานี": "Udon Thani", "อุตรดิตถ์": "Uttaradit",
    "อุทัยธานี": "Uthai Thani", "อุบลราชธานี": "Ubon Ratchathani",
}

# Abbreviations sometimes used in DDPM data
ABBREV_MAP = {
    "กรุงเทพ": "กรุงเทพมหานคร",
    "กทม": "กรุงเทพมหานคร",
    "กทม.": "กรุงเทพมหานคร",
    "อยุธยา": "พระนครศรีอยุธยา",
}

# All 77 province English names (for filling zeros)
ALL_PROVINCES_EN = set(PROVINCE_MAP.values())


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def _clean_prov(raw):
    """Clean province name: strip, handle abbreviations."""
    s = str(raw or "").strip()
    # Remove common prefixes like จ. or จังหวัด
    for prefix in ["จ.", "จังหวัด"]:
        if s.startswith(prefix):
            s = s[len(prefix):].strip()
    # Check abbreviations
    if s in ABBREV_MAP:
        s = ABBREV_MAP[s]
    return s


def _safe_int(val):
    """Convert a value to int, handling None, float, and string numbers."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().replace(",", "")
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


# ─────────────────────────────────────────────
# Parsers for each format
# ─────────────────────────────────────────────
def parse_2562(path):
    """Format 1: Summary by province -- Sheet "flood2562cut"
    Row 0: EN headers, Row 1: TH headers, Row 2+: data
    Col 0 = province (TH), Col 2 = number of tambons
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Try specific sheet name first, fall back to last sheet
    ws = wb["flood2562cut"] if "flood2562cut" in wb.sheetnames else wb[wb.sheetnames[-1]]
    result = {}
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip 2 header rows
        prov = _clean_prov(row[0])
        if not prov or prov not in PROVINCE_MAP:
            continue
        n_tambon = _safe_int(row[2])
        if n_tambon > 0:
            result[prov] = n_tambon
    wb.close()
    return result


def parse_granular(path, prov_col, tambon_col, skip_rows):
    """Format 2/3: Per-tambon/village rows -- group by province, count distinct tambons.
    Uses last sheet of the workbook.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[-1]]
    prov_tambons = {}  # { "province_th": {"tambon1", "tambon2", ...} }
    for row in ws.iter_rows(min_row=skip_rows + 1, values_only=True):
        if row is None:
            continue
        if len(row) <= max(prov_col, tambon_col):
            continue
        prov = _clean_prov(row[prov_col])
        tambon = str(row[tambon_col] or "").strip()
        # Remove common prefixes from tambon
        for prefix in ["ต.", "ตำบล"]:
            if tambon.startswith(prefix):
                tambon = tambon[len(prefix):].strip()
        if not prov or not tambon:
            continue
        # Validate province name -- must be in our map
        if prov not in PROVINCE_MAP:
            continue
        # Filter out long strings that are likely not province names
        if len(prov) > 30:
            continue
        if prov not in prov_tambons:
            prov_tambons[prov] = set()
        prov_tambons[prov].add(tambon)
    wb.close()
    return {p: len(t) for p, t in prov_tambons.items()}


def parse_2568(path):
    """Format 4: Summary by province -- Sheet "flood68"
    Row 0: EN headers, Row 1: TH headers, Row 2+: data
    Col 2 = province, Col 5 = number of tambons
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Try specific sheet name first, fall back to last sheet
    target = None
    for name in wb.sheetnames:
        if "68" in name or "อุทกภัย" in name:
            target = name
            break
    ws = wb[target] if target else wb[wb.sheetnames[-1]]
    result = {}
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip 2 header rows
        if row is None or len(row) <= 5:
            continue
        prov = _clean_prov(row[2])
        if not prov or prov not in PROVINCE_MAP:
            continue
        n_tambon = _safe_int(row[5])
        if n_tambon > 0:
            result[prov] = n_tambon
    wb.close()
    return result


# ─────────────────────────────────────────────
# Year -> parser mapping
# ─────────────────────────────────────────────
YEAR_PARSERS = {
    "2562": lambda: parse_2562(os.path.join(DDPM_DIR, "ddpm_flood_2562.xlsx")),
    "2563": lambda: parse_granular(os.path.join(DDPM_DIR, "ddpm_flood_2563.xlsx"), 0, 2, 1),   # no TH header row
    "2564": lambda: parse_granular(os.path.join(DDPM_DIR, "ddpm_flood_2564.xlsx"), 0, 2, 2),
    "2565": lambda: parse_granular(os.path.join(DDPM_DIR, "ddpm_flood_2565.xlsx"), 0, 2, 2),
    "2566": lambda: parse_granular(os.path.join(DDPM_DIR, "ddpm_flood_2566.xlsx"), 3, 7, 2),   # Col3=province, Col7=tambon
    "2567": lambda: parse_granular(os.path.join(DDPM_DIR, "ddpm_flood_2567.xlsx"), 5, 9, 2),   # Col5=province, Col9=tambon
    "2568": lambda: parse_2568(os.path.join(DDPM_DIR, "ddpm_flood_2568.xlsx")),
}


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("DDPM Flood Risk Builder")
    print(f"Source: {DDPM_DIR}")
    print(f"Output: {OUTPUT}")
    print("=" * 60)

    # 1. Parse each year
    all_years = {}
    for year, parser in YEAR_PARSERS.items():
        xlsx_path = os.path.join(DDPM_DIR, f"ddpm_flood_{year}.xlsx")
        if not os.path.exists(xlsx_path):
            print(f"  {year}: SKIP (file not found: {xlsx_path})")
            continue
        try:
            data = parser()
            all_years[year] = data
            print(f"  {year}: {len(data)} provinces")
        except Exception as e:
            print(f"  {year}: ERROR - {e}")

    if not all_years:
        print("\n[ERROR] No data parsed from any year!")
        sys.exit(1)

    n_years = len(all_years)
    print(f"\nParsed {n_years} years successfully.\n")

    # 2. Aggregate: for each province count years flooded + max tambons
    all_provs_th = set()
    for yr_data in all_years.values():
        all_provs_th.update(yr_data.keys())

    results = {}
    unmapped = []
    for prov_th in sorted(all_provs_th):
        prov_en = PROVINCE_MAP.get(prov_th)
        if not prov_en:
            unmapped.append(prov_th)
            continue
        years_flooded = sum(1 for yr_data in all_years.values() if prov_th in yr_data)
        max_tambons = max((yr_data.get(prov_th, 0) for yr_data in all_years.values()), default=0)
        pct = round(years_flooded / n_years * 100, 1)
        results[prov_en] = {
            "pct": pct,
            "flood_events": years_flooded,
            "total_years": n_years,
            "flood_tambons": max_tambons,
            "total_tambons": 0,
            "flood_rai": 0,
            "province_rai": 0,
        }

    if unmapped:
        print("Unmapped province names:")
        for u in unmapped:
            print(f"  WARNING: '{u}'")
        print()

    # 3. Fill provinces that never flooded with pct=0
    for prov_en in ALL_PROVINCES_EN:
        if prov_en not in results:
            results[prov_en] = {
                "pct": 0.0,
                "flood_events": 0,
                "total_years": n_years,
                "flood_tambons": 0,
                "total_tambons": 0,
                "flood_rai": 0,
                "province_rai": 0,
            }

    # 4. Print summary
    sorted_by_pct = sorted(results.items(), key=lambda x: (-x[1]["pct"], x[0]))

    print(f"Total provinces: {len(results)}")
    print(f"\n--- TOP 10 (highest flood frequency) ---")
    for prov, d in sorted_by_pct[:10]:
        print(f"  {prov:30s}  {d['pct']:5.1f}%  ({d['flood_events']}/{d['total_years']} yrs, max {d['flood_tambons']} tambons)")

    print(f"\n--- BOTTOM 10 (lowest flood frequency) ---")
    for prov, d in sorted_by_pct[-10:]:
        print(f"  {prov:30s}  {d['pct']:5.1f}%  ({d['flood_events']}/{d['total_years']} yrs, max {d['flood_tambons']} tambons)")

    # 5. Check specific provinces of interest (Andaman coast)
    check_provs = ["Phuket", "Phangnga", "Ranong", "Chumphon", "Krabi"]
    print(f"\n--- Andaman coast check ---")
    for prov in check_provs:
        d = results.get(prov, {})
        pct = d.get("pct", "N/A")
        events = d.get("flood_events", "N/A")
        tambons = d.get("flood_tambons", "N/A")
        print(f"  {prov:20s}  {pct}%  ({events}/{n_years} yrs, max {tambons} tambons)")

    # 6. Save flood-risk.json
    output_data = {
        "_meta": {
            "source": "กรมป้องกันและบรรเทาสาธารณภัย (ปภ.) / DDPM",
            "source_en": "Department of Disaster Prevention and Mitigation (DDPM)",
            "source_url": "https://catalog.disaster.go.th/dataset/dpm-gd027",
            "period": "2562-2568 (7 years / 2019-2025)",
            "method": "นับจากประกาศพื้นที่ภัยพิบัติอย่างเป็นทางการ — ครอบคลุมทุกประเภทน้ำท่วม",
            "updated": date.today().isoformat(),
            "note": "% ปีที่ท่วมจาก 7 ปีล่าสุด · ข้อมูลจากประกาศเขตภัยพิบัติ ปภ.",
        },
        "provinces": dict(sorted(results.items())),
    }

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    print(f"\n[SAVED] {OUTPUT}")
    print(f"  {len(results)} provinces | {n_years} years")


if __name__ == "__main__":
    main()
